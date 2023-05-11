import asyncio
import copy
import json
import os
import traceback
from contextlib import asynccontextmanager
from typing import List

import pandas as pd
import tqdm
from openpyxl import load_workbook

from superai.llm.actions import GoogleSearchAction
from superai.llm.data_types.message import ChatMessage
from superai.llm.foundation_models.openai import ChatGPT
from superai.llm.logger import logger
from superai.llm.utilities.parser_utils import Parser

# Read input data from CSV file
data_path = os.path.abspath("./examples/local_input_data/uc2.csv")
input_data = pd.read_csv(data_path)
input_data = input_data.where(pd.notnull(input_data), None).to_dict(orient="records")

# Create a global semaphore
semaphore = asyncio.Semaphore(50)  # Adjust the value to the desired concurrency limit

prompt_template = """You are Business Detective, You need to find the real name, URL, and phone number of a business given a name and address from a credit card statement and google search results.

Master Goal:
The user has inputted: 
{user_input}

You need take that input and turn it into a valid JSON that can be parsed by and turn it into and turn it into a valid JSON output of the form (ensure the response can be parsed by Python json.loads):
{{'name': {{'certainty':"number 0-1", 'value':'string', 'source':"url of text it came from"}}, 'address': {{'certainty':"number 0-1", 'value':'string', 'source':"url of text it came from"}}, 'corporate_name':{{'certainty':"number 0-1", 'value':'string', 'source':"url of text it came from"}},  'corporate_address':{{'certainty':"number 0-1", 'value':'string', 'source':"url of text it came from"}}, 'phone': {{'certainty':"number 0-1", 'value':'string', 'source':"url of text it came from"}}, 'url': {{'certainty':"number 0-1", 'value':'string', 'source':"url of text it came from"}}
Only respond with the JSON output nothing more. 

Constraints: 
- Names should be only the business names not anything extra
- The domains should be the root domain with no extra sites and not particular locations 
- For each field also add a properly calibrated certainty score between 0 and 1 (e.g. 0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0). Where for example 0 means you have 0 evidence, that the probability is 0% that the answer is correct, .5 means you think there is a 50% probability of the answer being correct, and 1.0 is there is a 100% probability of the answer being correct. Always give an answer.
- For the source give a specific URL, if there is not URL then give a segment of text where you extracted it from. Always be specific.

Google search results: 
{search_results}
"""


@asynccontextmanager
async def semaphore_context(sem):
    await sem.acquire()
    try:
        yield
    finally:
        sem.release()


def get_user_feedback(output_dict, user_input):
    if output_dict["fields_to_check"]:
        logger.debug(f"Here is the user's input:")
        for key, value in user_input.items():
            logger.debug(f"- {key}: {value}")
        logger.debug(f"\n\nOutput of the AI:")
        for key, value in output_dict["output"].items():
            logger.debug(f"- {key}: {value['value']} ({value['certainty']*100}% certainty)")
        logger.debug("\nPlease check the following fields")
        output_dict_copy = copy.deepcopy(output_dict)
        for key, value in output_dict_copy["fields_to_check"].items():
            while True:
                pick = input(
                    f"Is '{value['value']}' the {key} of '{user_input['Brand Name']}'?\n(AI has certaintly {value['certainty']})\n Link: {output_dict['output'][key]['source']}? (y,n)"
                ).lower()
                if pick == "y":
                    logger.debug("Confirmed... thanks. ")
                    output_dict["fields_to_check"][key]["user_feedback"] = {"selection": pick, "user_feedback": ""}
                    break
                elif pick == "n":
                    user_feedback = input(f"What is the correct {key}?")
                    output_dict["fields_to_check"][key]["user_feedback"] = {
                        "selection": pick,
                        "user_feedback": user_feedback,
                    }
                    logger.debug(f"Updated {key} to {user_feedback}")
                    break
                else:
                    logger.debug("You have to choose y or n")
    return output_dict


def create_search_string(user_input):
    search_string = "What is the brand name, address, corporate name, corporate address, phone number, and url of: "
    search_string += f"{user_input['Brand Name']}"
    if user_input["url"]:
        search_string += f" {user_input['url']}"
    if user_input["phone"]:
        search_string += f" {user_input['phone']}"
    # if user_input['gmr_vmid_name']:
    #     search_string += f" {user_input['gmr_vmid_name']}"
    return search_string


def parse_search_results(search_results: List[str]):
    search_results = search_results.replace("\n", "").lower()
    return search_results


def write_to_file(data, file_path):
    # Creating the dataframes
    outputs_df = pd.DataFrame(data["outputs"])
    fields_to_check_df = pd.DataFrame(data["fields_to_check"])

    # Creating super columns
    columns = ["name", "address", "corporate_name", "corporate_address", "phone", "url", "address"]

    for col in columns:
        for df in [outputs_df, fields_to_check_df]:
            df[f"{col}_certainty"] = ""
            df[f"{col}_value"] = ""
            df[f"{col}_source"] = ""

            # Rearranging the columns
            cols = df.columns.tolist()
            new_cols = cols[:-3] + [cols[-3]] + [cols[-1]] + [cols[-2]]
            df = df[new_cols]

    if os.path.exists(file_path):
        # Read existing Excel file
        book = load_workbook(file_path)
        writer = pd.ExcelWriter(file_path, engine="openpyxl")
        writer.book = book

        # Append new data to existing sheets
        for sheet_name, df in [("Outputs", outputs_df), ("Fields to Check", fields_to_check_df)]:
            if sheet_name in book.sheetnames:
                sheet_df = pd.read_excel(file_path, sheet_name=sheet_name)
                sheet_df = sheet_df.append(df, ignore_index=True)
            else:
                sheet_df = df

            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    else:
        # Create a new Excel file
        writer = pd.ExcelWriter(file_path, engine="openpyxl")
        outputs_df.to_excel(writer, sheet_name="Outputs", index=False)
        fields_to_check_df.to_excel(writer, sheet_name="Fields to Check", index=False)

    writer.save()


async def get_output(user_input, prompt_template, human_feedback_certainty=1.0, iteration=None):
    async with semaphore_context(semaphore):
        try:
            # Create search results
            search_tool = GoogleSearchAction(n_results=10)
            search_results = search_tool.run(input=create_search_string(user_input))
            parsed_search_results = parse_search_results(search_results)

            if parsed_search_results == "no good google search result was found":
                output_logs_path = os.path.abspath("./examples/local_output_data/errors.log")
                with open(output_logs_path, "a") as f:
                    f.write(
                        json.dumps(
                            f"Error: No good google search result was found for {create_search_string(user_input)}"
                        )
                    )
                    f.write("\n")
                search_results = search_tool.run(input=user_input["Brand Name"])
                parsed_search_results = parse_search_results(search_results)
                if parsed_search_results == "no good google search result was found":
                    output_logs_path = os.path.abspath("./examples/local_output_data/errors.log")
                    with open(output_logs_path, "a") as f:
                        f.write(
                            json.dumps(f"Error: No good google search result was found for {user_input['Brand Name']}")
                        )
                        f.write("\n")
                    return None

            # Log search results summary
            logger.debug(f"Iteration {iteration}")

            # Create the prompt
            prompt = prompt_template.format(user_input=user_input, search_results=parsed_search_results)
            logger.debug(f"Prompt: {prompt}")

            # Send to gpt-3.5
            chat_gpt = ChatGPT(engine="gpt-3.5-turbo", temperature=0)
            response = chat_gpt.predict(input=ChatMessage(role="system", content=prompt))

            # Parse response
            output_parser = Parser(use_ai=False)
            valid_response = output_parser.to_dict(response)
            logger.debug(f"Search results: {str(valid_response)}")

            output_dict = {"output": valid_response}
            fields_to_check = {}
            fields_above_threshold = 0
            fields_below_threshold = 0
            add_to_review = False
            for key, value in valid_response.items():
                certainty = float(value["certainty"])
                if certainty < human_feedback_certainty:
                    fields_to_check[key] = value
                    fields_below_threshold += 1
                    add_to_review = True
                else:
                    fields_above_threshold += 1

            output_dict["fields_to_check"] = fields_to_check

            if add_to_review:
                output_file_path = os.path.abspath("./examples/local_output_data/fields_to_check.xls")
                write_to_file(output_dict, output_file_path)
                add_to_review = False

            # Log fields above and below human_feedback_certainty
            logger.debug(
                f"Iteration {iteration}: {fields_above_threshold} fields above threshold, {fields_below_threshold} fields below threshold"
            )

            # # Get user feedback
            # updated_output_dict = get_user_feedback(output_dict=output_dict, user_input=user_input)

            # Log output summary
            # logger.debug(f"Iteration {iteration}: Output summary:")
            # for key, value in updated_output_dict['output'].items():
            #     logger.debug(f"{key}: {value['value']} ({value['certainty'] * 100}% certainty)")

            # Write the result to a file
            output_file_path = os.path.abspath("./examples/local_output_data/output_results.xls")
            write_to_file(output_dict, output_file_path)

            return output_dict

        except Exception as e:
            error_message = f"Error: {str(e)}\nTraceback: {traceback.format_exc()}"
            output_logs_path = os.path.abspath("./examples/local_output_data/failed_runs.log")
            with open(output_logs_path, "a") as f:
                f.write(json.dumps(error_message))
                f.write("\n")

            return None


async def main():
    tasks = [get_output(row, prompt_template, iteration=i) for i, row in enumerate(input_data[:2])]
    for task in tqdm.tqdm(asyncio.as_completed(tasks), total=len(tasks), desc="Processing tasks"):
        await task


if __name__ == "__main__":
    asyncio.run(main())
